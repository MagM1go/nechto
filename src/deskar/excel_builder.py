import logging
from pathlib import Path
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from deskar.models import (
    EXCEL_COLUMN_WIDTHS,
    EXCEL_HEADERS,
    CatalogRecord,
)

logger = logging.getLogger(__name__)

HEADER_FILL_COLOR: Final[str] = "2F4F4F"
EVEN_ROW_FILL_COLOR: Final[str] = "F0F0F0"


class ExcelBuilder:
    def __init__(self, output_path: Path) -> None:
        self._output_path = output_path
        self._workbook: openpyxl.Workbook | None = None
        self._worksheet: Worksheet | None = None
        self._current_row = 1

    def _get_or_create_workbook(self) -> openpyxl.Workbook:
        if self._workbook is None:
            self._workbook = openpyxl.Workbook()
            self._worksheet = self._workbook.active
            if self._worksheet:
                self._worksheet.title = "Catalog"
        return self._workbook

    def _get_worksheet(
        self,
    ) -> Worksheet:
        workbook = self._get_or_create_workbook()
        return workbook.active  # pyright: ignore[reportReturnType]

    def add_headers(self) -> None:
        ws = self._get_worksheet()

        header_fill = PatternFill(
            start_color=HEADER_FILL_COLOR,
            end_color=HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(
            bold=True,
            color="FFFFFF",
            name="Arial",
            size=10,
        )
        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        ws.row_dimensions[1].height = 25
        self._current_row = 2

    def add_record(self, record: CatalogRecord) -> None:
        ws = self._get_worksheet()

        data_font = Font(name="Arial", size=9)
        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        is_even_row = self._current_row % 2 == 0
        row_fill = (
            PatternFill(
                start_color=EVEN_ROW_FILL_COLOR,
                end_color=EVEN_ROW_FILL_COLOR,
                fill_type="solid",
            )
            if is_even_row
            else None
        )

        row_values = record.to_excel_row()

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(
                row=self._current_row,
                column=col,
                value=value,
            )
            cell.font = data_font
            cell.alignment = center_alignment
            if row_fill:
                cell.fill = row_fill

        self._current_row += 1

    def add_records(self, records: list[CatalogRecord]) -> None:
        for record in records:
            self.add_record(record)

    def apply_formatting(self) -> None:
        ws = self._get_worksheet()

        for col, width in enumerate(EXCEL_COLUMN_WIDTHS, start=1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A2"

    def save(self) -> None:
        workbook = self._get_or_create_workbook()

        self._output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook.save(self._output_path)
        logger.info(f"Saved Excel file: {self._output_path}")

    def build_and_save(self, records: list[CatalogRecord]) -> int:
        self.add_headers()
        self.add_records(records)
        self.apply_formatting()
        self.save()

        return len(records)

    def get_record_count(self) -> int:
        return self._current_row - 2

    @property
    def output_path(self) -> Path:
        return self._output_path


def build_excel_from_records(
    records: list[CatalogRecord],
    output_path: Path,
) -> int:
    builder = ExcelBuilder(output_path)
    return builder.build_and_save(records)
